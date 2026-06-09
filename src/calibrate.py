"""校准模块：对比人工评分与 Judge 模型评分。

从 run 目录下的 analysis_by_human.xlsx 读取人工评估结果，
与 scores.jsonl 中的 Judge 评分并排展示，追加 Calibration 页到 report.xlsx。
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.constants import RESULTS_DIR
from src.reporter import load_scores, _short_report_name

HUMAN_FILE = "analysis_by_human.xlsx"

DIM_NAME_MAP = {
    "relevance": "相关性",
    "factuality": "事实性",
    "fluency": "流畅性",
    "structure": "结构化",
    "timeliness": "实时性",
    "localization": "本地化",
    "search_planning": "搜索规划",
    "search_results": "搜索结果",
    "search_quality": "搜索质量",
    "overall": "综合",
}


def load_human_scores(excel_path: Path) -> dict[int, dict]:
    """从 analysis_by_human.xlsx 读取人工评估结果。

    Returns:
        dict keyed by row_index, value = {row_index, query, analysis_by_human}
    """
    try:
        df = pd.read_excel(excel_path)
    except ValueError:
        raise FileNotFoundError(
            f"无法读取人工评估文件: {excel_path}"
        )

    result: dict[int, dict] = {}
    for _, row in df.iterrows():
        ri = int(row["row_index"])
        result[ri] = {
            "row_index": ri,
            "query": str(row.get("query", "")),
            "analysis_by_human": str(row.get("analysis_by_human", "")),
        }
    return result


def generate_calibration_report(run_name: str) -> Path:
    """生成校准报告，在 report.xlsx 中追加 Calibration 页。

    人工评估从 results/<run_name>/analysis_by_human.xlsx 读取，
    Judge 评分从 scores.jsonl 读取。

    Returns:
        更新后的 Excel 文件路径
    """
    result_dir = RESULTS_DIR / run_name

    # 人工评估文件
    human_path = result_dir / HUMAN_FILE
    if not human_path.exists():
        raise FileNotFoundError(
            f"人工评估文件不存在: {human_path}\n"
            f"  请在 {result_dir} 下创建 {HUMAN_FILE}，包含 row_index、query、analysis_by_human 列"
        )

    # Excel 报告
    target_excel = result_dir / f"{_short_report_name(run_name)}_report.xlsx"
    if not target_excel.exists():
        raise FileNotFoundError(f"Excel 报告不存在: {target_excel}\n  请先运行 eval 或 export 生成报告。")

    # Judge 评分
    scores_path = result_dir / "scores.jsonl"
    judge_scores = load_scores(scores_path)

    # 人工评估
    human_scores = load_human_scores(human_path)

    dimensions = _detect_dimensions(judge_scores)
    dim_cols = [d for d in dimensions if d not in ("analysis", "overall")]
    columns = ["row_index", "query", "judge_overall"] + dim_cols + ["human_analysis"]

    all_row_indices = sorted(set(list(judge_scores.keys()) + list(human_scores.keys())))
    matched = 0
    rows_data = []
    for ri in all_row_indices:
        judge = judge_scores.get(ri, {})
        human = human_scores.get(ri, {})
        if judge and human:
            matched += 1

        row = {"row_index": ri, "query": _truncate(human.get("query", judge.get("query", "")))}
        row["judge_overall"] = judge.get("overall")
        for d in dim_cols:
            row[d] = judge.get(d)
        row["human_analysis"] = human.get("analysis_by_human", "")
        rows_data.append(row)

    df_cal = pd.DataFrame(rows_data, columns=columns)

    wb = load_workbook(target_excel)
    if "Calibration" in wb.sheetnames:
        del wb["Calibration"]

    ws = wb.create_sheet("Calibration")

    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(rows_data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    for i, _ in enumerate(dim_cols, 4):
        ws.column_dimensions[get_column_letter(i)].width = 10
    human_col = len(columns)
    ws.column_dimensions[get_column_letter(human_col)].width = 60

    wb.save(target_excel)

    _print_summary(judge_scores, human_scores, matched, len(all_row_indices))

    return target_excel


def _detect_dimensions(scores: dict[int, dict]) -> list[str]:
    for s in scores.values():
        dims = [k for k in s if k not in ("row_index", "query", "response_summary", "error")]
        if dims:
            return dims
    return []


def _truncate(text: str, max_len: int = 120) -> str:
    if len(text) <= max_len:
        return text
    return text[:max_len] + "..."


def _print_summary(
    judge_scores: dict[int, dict],
    human_scores: dict[int, dict],
    matched: int,
    total: int,
):
    print(f"[calibrate] 人工评估行数: {len(human_scores)}")
    print(f"[calibrate] Judge 评分行数: {len(judge_scores)}")
    print(f"[calibrate] 匹配行数: {matched}/{total}")

    overalls = [s["overall"] for s in judge_scores.values() if "overall" in s and s["overall"] is not None]
    if overalls:
        avg_overall = sum(overalls) / len(overalls)
        print(f"[calibrate] Judge 平均综合分: {avg_overall:.2f}")

    if judge_scores and human_scores:
        dims = _detect_dimensions(judge_scores)
        for dim in dims:
            if dim == "overall":
                continue
            j_vals = [s[dim] for s in judge_scores.values() if dim in s and isinstance(s[dim], (int, float))]
            if j_vals:
                avg = sum(j_vals) / len(j_vals)
                if avg >= 4.8:
                    chinese = DIM_NAME_MAP.get(dim, dim)
                    print(f"[calibrate] [WARN] {chinese} ({dim}): Judge 均分 {avg:.2f}，可能存在打分偏高")
